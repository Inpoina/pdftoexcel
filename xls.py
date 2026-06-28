#!/usr/bin/env python3
"""
PDF Tabel Scan → Excel Converter (Multi-PDF) v9
Mengkonversi satu atau lebih PDF hasil scan (berisi dua kolom: ID, Nilai)
menjadi SATU file Excel dengan:
  • Sheet "GABUNGAN"  → semua data dari semua PDF, dengan kolom Sumber PDF
  • Sheet per PDF     → data masing-masing file secara terpisah
  • Verifikasi PLU 8 digit ke database MySQL (tve8.produk, kolom plu)
    - Baris PLU "TIDAK DITEMUKAN" di DB otomatis DIHAPUS dari output Excel
    - Hanya PLU yang ada di database yang masuk ke Excel

Cara pakai:
    python v9.py file1.pdf file2.pdf file3.pdf
    python v9.py *.pdf
    python v9.py *.pdf -o hasil.xlsx
    python v9.py *.pdf --dpi 400          # scan buram
    python v9.py *.pdf --no-db            # lewati verifikasi MySQL (semua baris masuk)
"""

import sys
import re
import os
import glob
import subprocess
import argparse
import statistics
from pathlib import Path

# ---------------------------------------------------------------------------
# Dependency check & install
# ---------------------------------------------------------------------------
def check_and_install(packages):
    import importlib
    for pkg, import_name in packages:
        try:
            importlib.import_module(import_name)
        except ImportError:
            print(f"[INFO] Menginstall {pkg}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q",
                                   "--break-system-packages"])

check_and_install([
    ("pytesseract", "pytesseract"),
    ("Pillow", "PIL"),
    ("openpyxl", "openpyxl"),
    ("pdf2image", "pdf2image"),
    ("mysql-connector-python", "mysql.connector"),
])

import pytesseract
from PIL import Image, ImageFilter, ImageEnhance
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import mysql.connector


# ---------------------------------------------------------------------------
# Konfigurasi MySQL  ← sesuaikan jika berbeda
# ---------------------------------------------------------------------------
DB_CONFIG = {
    "host":     "localhost",
    "user":     "root",
    "password": "",          # tanpa password
    "database": "tve8",
    "connection_timeout": 5,
}
DB_TABLE  = "produk"
DB_COLUMN = "plu"           # kolom PLU 8 digit


# ---------------------------------------------------------------------------
# Verifikasi PLU ke MySQL
# ---------------------------------------------------------------------------
def load_plu_set() -> set | None:
    """
    Muat semua nilai PLU dari tabel ke dalam set Python.
    Lebih efisien dari query per-baris.
    Kembalikan None jika koneksi gagal.
    """
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        cur  = conn.cursor()
        cur.execute(f"SELECT `{DB_COLUMN}` FROM `{DB_TABLE}`")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        plu_set = {str(r[0]).strip().zfill(8) for r in rows}
        print(f"[DB] Terhubung ke {DB_CONFIG['database']}.{DB_TABLE} "
              f"— {len(plu_set)} PLU dimuat.")
        return plu_set
    except mysql.connector.Error as e:
        print(f"[DB ERROR] Tidak dapat terhubung ke MySQL: {e}")
        return None


def verify_plu(id_val: str, plu_set: set | None) -> str:
    """
    Kembalikan status verifikasi PLU:
      'VALID'           — PLU ditemukan di database
      'TIDAK DITEMUKAN' — PLU tidak ada di database
      'DB ERROR'        — koneksi/muat DB gagal
    """
    if plu_set is None:
        return "DB ERROR"
    return "VALID" if id_val in plu_set else "TIDAK DITEMUKAN"




# ---------------------------------------------------------------------------
# OCR helpers  (tidak berubah dari v4)
# ---------------------------------------------------------------------------
def preprocess(image: Image.Image) -> Image.Image:
    img = image.convert("L")
    img = img.filter(ImageFilter.SHARPEN)
    img = ImageEnhance.Contrast(img).enhance(2.0)
    return img


def binarize(image: Image.Image, threshold: int = 150) -> Image.Image:
    img = image.convert("L")
    img = img.point(lambda x: 0 if x < threshold else 255, "1")
    return img


def pdf_to_images(pdf_path: str, dpi: int = 300) -> list:
    print(f"  [OCR] Membaca PDF: {pdf_path}")
    images = convert_from_path(pdf_path, dpi=dpi)
    print(f"        {len(images)} halaman ditemukan.")
    return images


def parse_line_tokens(toks: list) -> tuple | None:
    # Bersihkan noise per-token: hapus karakter | _ - yang sering muncul
    # dari garis tabel OCR, lalu filter token yang jadi kosong.
    # Ini menangani kasus seperti: ['|', '20096590', '|', '8']
    # atau ['_20141070', '|', '1'] yang gagal parse sebelumnya.
    clean = [re.sub(r"[|_\-]", "", t).strip() for t in toks]
    clean = [t for t in clean if t]

    if len(clean) < 2:
        return None
    id_tok = clean[0]
    if not re.fullmatch(r"\d{8}", id_tok):
        return None
    val_tok = clean[1].upper()
    if re.fullmatch(r"#?N/?A", val_tok) or val_tok in ("#N/A", "NA", "#NA", "N/A"):
        return (id_tok, "#N/A")
    if re.fullmatch(r"\d+", val_tok):
        return (id_tok, int(val_tok))
    return None


def parse_data_with_position(data: dict) -> list[tuple]:
    lines = {}
    n = len(data["text"])
    for i in range(n):
        text = data["text"][i].strip()
        if not text:
            continue
        key = (data["block_num"][i], data["par_num"][i], data["line_num"][i])
        lines.setdefault(key, []).append((data["left"][i], data["top"][i], text))

    rows = []
    for key, words in lines.items():
        words.sort(key=lambda w: w[0])
        tokens_raw = " ".join(w[2] for w in words)
        tokens_raw = tokens_raw.replace("|", " ").replace("—", "").replace("_", "")
        toks = tokens_raw.split()
        parsed = parse_line_tokens(toks)
        if parsed is None:
            continue
        y_avg = sum(w[1] for w in words) / len(words)
        rows.append((parsed[0], parsed[1], y_avg))
    return rows


def ocr_page(image: Image.Image) -> list[tuple]:
    candidates = []

    data1 = pytesseract.image_to_data(image, config="--psm 6", output_type=pytesseract.Output.DICT)
    candidates.append(parse_data_with_position(data1))

    pre = preprocess(image)
    data2 = pytesseract.image_to_data(pre, config="--psm 6", output_type=pytesseract.Output.DICT)
    candidates.append(parse_data_with_position(data2))

    data3 = pytesseract.image_to_data(pre, config="--psm 4", output_type=pytesseract.Output.DICT)
    candidates.append(parse_data_with_position(data3))

    bw = binarize(image)
    data4 = pytesseract.image_to_data(bw, config="--psm 6", output_type=pytesseract.Output.DICT)
    candidates.append(parse_data_with_position(data4))

    vote_count = {}
    id_seen_in = {}
    y_positions = {}
    for rows in candidates:
        for id_tok, val, y in rows:
            pair = (id_tok, val)
            vote_count[pair] = vote_count.get(pair, 0) + 1
            id_seen_in.setdefault(id_tok, set()).add(pair)
            y_positions.setdefault(id_tok, []).append(y)

    seen_ids = set()
    result = []
    for rows in candidates:
        for id_tok, val, y in rows:
            if id_tok in seen_ids:
                continue
            seen_ids.add(id_tok)
            variants = id_seen_in[id_tok]
            best_pair = max(variants, key=lambda p: vote_count[p])
            total_votes_for_id = sum(vote_count[p] for p in variants)
            low_confidence = total_votes_for_id <= 1
            y_median = statistics.median(y_positions[id_tok])
            result.append((best_pair[0], best_pair[1], low_confidence, y_median))

    result.sort(key=lambda r: r[3])
    return [(id_tok, val, low_conf) for id_tok, val, low_conf, _ in result]


def ocr_pdf(pdf_path: str, dpi: int = 300) -> list[tuple]:
    """Jalankan OCR seluruh halaman satu PDF, kembalikan list (id, nilai, low_conf)."""
    images = pdf_to_images(pdf_path, dpi=dpi)
    all_rows = []
    for idx, img in enumerate(images, start=1):
        print(f"        Halaman {idx}/{len(images)}...", end=" ", flush=True)
        rows = ocr_page(img)
        print(f"{len(rows)} baris.")
        all_rows.extend(rows)
    return all_rows


# ---------------------------------------------------------------------------
# Excel styles
# ---------------------------------------------------------------------------
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
NA_FILL      = PatternFill("solid", fgColor="FCE4D6")
SUM_FILL     = PatternFill("solid", fgColor="E2EFDA")
COMBINE_FILL = PatternFill("solid", fgColor="2E4057")  # header sheet gabungan (biru tua)
THIN         = Side(style="thin", color="BFBFBF")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header_cell(cell, fg_fill, font_color="FFFFFF"):
    cell.font      = Font(name="Arial", bold=True, color=font_color, size=11)
    cell.fill      = fg_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = BORDER


def _style_data_cell(cell, fill=None):
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = BORDER
    if fill:
        cell.fill = fill


def _write_summary(ws, data_end: int, sr: int, has_source_col: bool = False):
    """Tulis blok ringkasan di bawah tabel data."""
    # Sheet per-PDF : No. | ID | Nilai  → id=B, val=C
    # Sheet gabungan: No. | Sumber | ID | Nilai → id=C, val=D
    id_col  = "C" if has_source_col else "B"
    val_col = "D" if has_source_col else "C"

    labels = ["Total Data", "Data #N/A", "Data Valid"]
    formulas = [
        f"=COUNTA({id_col}2:{id_col}{data_end})",
        f'=COUNTIF({val_col}2:{val_col}{data_end},"#N/A")',
        f"=B{sr}-B{sr+1}" if not has_source_col else f"=C{sr}-C{sr+1}",
    ]
    for offset, (lbl, frm) in enumerate(zip(labels, formulas)):
        r = sr + offset
        ws.cell(row=r, column=1, value=lbl)
        ws.cell(row=r, column=2, value=frm)
        ws.cell(row=r, column=3, value="baris")
        for col in range(1, 4):
            cell = ws.cell(row=r, column=col)
            cell.font      = Font(name="Arial", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center")
            cell.border    = BORDER
            cell.fill      = SUM_FILL


def write_per_pdf_sheet(ws, rows: list[tuple], pdf_name: str):
    """
    Isi satu sheet untuk satu PDF.
    Kolom: No. | ID (PLU) | Nilai
    rows: list of (id_val, nilai)  — sudah difilter hanya VALID
    """
    headers = ["No.", "ID (PLU)", "Nilai"]
    for col, hdr in enumerate(headers, start=1):
        _style_header_cell(ws.cell(row=1, column=col, value=hdr), HEADER_FILL)
    ws.row_dimensions[1].height = 22

    for i, (id_val, nilai) in enumerate(rows, start=1):
        row_num = i + 1
        is_na   = (nilai == "#N/A")
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=id_val)
        ws.cell(row=row_num, column=3, value=nilai)
        row_fill = NA_FILL if is_na else (ALT_FILL if i % 2 == 0 else None)
        for col in range(1, 4):
            _style_data_cell(ws.cell(row=row_num, column=col), row_fill)

    total    = len(rows)
    data_end = total + 1
    sr       = total + 3
    _write_summary(ws, data_end, sr, has_source_col=False)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A2"


def write_combined_sheet(ws, combined: list[tuple]):
    """
    Sheet GABUNGAN.
    Kolom: No. | Sumber PDF | ID (PLU) | Nilai
    combined: list of (pdf_name, id_val, nilai)  — sudah difilter hanya VALID
    """
    headers = ["No.", "Sumber PDF", "ID (PLU)", "Nilai"]
    for col, hdr in enumerate(headers, start=1):
        _style_header_cell(ws.cell(row=1, column=col, value=hdr), COMBINE_FILL)
    ws.row_dimensions[1].height = 22

    for i, (pdf_name, id_val, nilai) in enumerate(combined, start=1):
        row_num = i + 1
        is_na   = (nilai == "#N/A")
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=pdf_name)
        ws.cell(row=row_num, column=3, value=id_val)
        ws.cell(row=row_num, column=4, value=nilai)
        row_fill = NA_FILL if is_na else (ALT_FILL if i % 2 == 0 else None)
        for col in range(1, 5):
            _style_data_cell(ws.cell(row=row_num, column=col), row_fill)

    total    = len(combined)
    data_end = total + 1
    sr       = total + 3
    _write_summary(ws, data_end, sr, has_source_col=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A2"


def safe_sheet_name(name: str, max_len: int = 31) -> str:
    """Excel membatasi nama sheet 31 karakter, tanpa karakter khusus."""
    for ch in r'\/*?[]':
        name = name.replace(ch, "_")
    return name[:max_len]


# ---------------------------------------------------------------------------
# KONFIGURASI PATH  ← ubah di sini sesuai kebutuhan
# ---------------------------------------------------------------------------

# Folder tempat file-file PDF berada.
# Contoh Windows : r"C:\Users\Anda\Dokumen\PDF"
# Contoh Linux   : "/home/anda/dokumen/pdf"
# None           : tidak dipakai, path diambil dari argumen CLI saja
PDF_SOURCE_DIR = "~/storage/shared/python/"

# Path lengkap file Excel hasil output.
# Contoh Windows : r"C:\Users\Anda\Dokumen\hasil\gabungan.xlsx"
# Contoh Linux   : "/home/anda/dokumen/hasil/gabungan.xlsx"
# None           : pakai argumen -o dari CLI, atau default "gabungan.xlsx"
OUTPUT_PATH = "~/storage/shared/python/gabung.xlsx"

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Konversi beberapa PDF scan tabel → satu file Excel (multi-sheet)"
    )
    parser.add_argument(
        "inputs", nargs="*",
        help=(
            "Nama file PDF yang ingin diproses (boleh lebih dari satu, boleh wildcard). "
            "Jika tidak diisi, semua *.pdf di PDF_SOURCE_DIR diproses. "
            "Jika PDF_SOURCE_DIR = None, dicari di folder kerja saat ini."
        )
    )
    parser.add_argument(
        "-o", "--output", default=None,
        help="Nama file Excel output (default: gabungan.xlsx)"
    )
    parser.add_argument(
        "--dpi", type=int, default=300,
        help="Resolusi OCR (default: 300). Naikkan ke 400 untuk scan buram."
    )
    parser.add_argument(
        "--no-db", action="store_true",
        help="Lewati verifikasi PLU ke MySQL."
    )
    args = parser.parse_args()

    # ------------------------------------------------------------------
    # Tentukan daftar pola PDF yang akan diproses.
    #
    # Aturan resolusi path:
    #   1. Tidak ada argumen  -> semua *.pdf di src_dir (PDF_SOURCE_DIR atau cwd)
    #   2. Nama file saja (mis. "laporan.pdf") -> digabung dengan src_dir
    #   3. Path lengkap / wildcard dengan folder -> pakai apa adanya
    # ------------------------------------------------------------------
    src_dir = Path(PDF_SOURCE_DIR).expanduser().resolve() if PDF_SOURCE_DIR else Path.cwd()

    if PDF_SOURCE_DIR and not src_dir.is_dir():
        print(f"[ERROR] PDF_SOURCE_DIR tidak ditemukan: {PDF_SOURCE_DIR}")
        sys.exit(1)

    if not args.inputs:
        # Tanpa argumen -> ambil semua PDF di src_dir
        input_patterns = [str(src_dir / "*.pdf")]
        print(f"[INFO] Tidak ada argumen — memproses semua PDF di: {src_dir}")
    else:
        # Ada argumen: resolusi path per token
        input_patterns = []
        for token in args.inputs:
            p = Path(token)
            if p.is_absolute() or p.parent != Path("."):
                # Path lengkap atau sudah punya komponen folder -> pakai apa adanya
                input_patterns.append(token)
            else:
                # Nama file / wildcard saja -> gabungkan dengan src_dir
                input_patterns.append(str(src_dir / token))

    # Ekspansi wildcard manual (berguna di Windows yang tidak expand glob)
    pdf_files = []
    for pattern in input_patterns:
        expanded = glob.glob(pattern)
        if expanded:
            pdf_files.extend(expanded)
        elif os.path.exists(pattern):
            pdf_files.append(pattern)
        else:
            print(f"[WARN] Tidak ditemukan: {pattern}")

    pdf_files = sorted(set(pdf_files))  # deduplikasi & urutkan

    if not pdf_files:
        print("[ERROR] Tidak ada file PDF yang ditemukan.")
        sys.exit(1)

    # ------------------------------------------------------------------
    # Tentukan path output.
    # Prioritas: OUTPUT_PATH (variable) → argumen -o CLI → default
    # ------------------------------------------------------------------
    if OUTPUT_PATH:
        output_path = str(Path(OUTPUT_PATH).expanduser().resolve())
        out_dir = Path(output_path).parent
        out_dir.mkdir(parents=True, exist_ok=True)
        print(f"[INFO] Menggunakan OUTPUT_PATH: {output_path}")
    else:
        output_path = args.output or "gabungan.xlsx"

    print(f"{'='*60}")
    print(f"  PDF ditemukan : {len(pdf_files)} file")
    print(f"  Output        : {output_path}")
    print(f"  DPI           : {args.dpi}")
    print(f"{'='*60}")

    # ------------------------------------------------------------------
    # Muat PLU dari MySQL sekali sebelum loop (lebih efisien)
    # ------------------------------------------------------------------
    if args.no_db:
        plu_set = None
        print("[DB] Verifikasi PLU dilewati (--no-db).")
    else:
        plu_set = load_plu_set()
        if plu_set is None:
            print("[DB] Verifikasi PLU dinonaktifkan karena koneksi gagal.")
            print("     Kolom 'Status PLU' akan berisi 'DB ERROR'.")


    wb = Workbook()
    # Hapus sheet default yang dibuat otomatis
    wb.remove(wb.active)

    # Sheet GABUNGAN dibuat duluan agar tampil paling kiri
    ws_combined = wb.create_sheet("GABUNGAN")
    combined_rows = []   # list of (pdf_name, id_val, nilai)

    total_rows_all = 0

    for pdf_idx, pdf_path in enumerate(pdf_files, start=1):
        pdf_name = Path(pdf_path).stem  # nama file tanpa ekstensi

        print(f"\n[{pdf_idx}/{len(pdf_files)}] Memproses: {pdf_path}")

        if not os.path.exists(pdf_path):
            print(f"  [ERROR] File tidak ditemukan, dilewati.")
            continue

        rows = ocr_pdf(pdf_path, dpi=args.dpi)

        if not rows:
            print(f"  [WARN] Tidak ada data diekstrak dari {pdf_path}, sheet tetap dibuat (kosong).")

        # Verifikasi PLU dan filter: buang yang TIDAK DITEMUKAN
        rows_verified = []
        n_dibuang = 0
        for id_val, nilai, low_conf in rows:
            status = verify_plu(id_val, plu_set)
            if status == "TIDAK DITEMUKAN":
                n_dibuang += 1
                continue  # lewati — tidak masuk Excel
            rows_verified.append((id_val, nilai))

        # Buat sheet per PDF
        sheet_name = safe_sheet_name(pdf_name)
        # Hindari nama sheet duplikat
        existing = [ws.title for ws in wb.worksheets]
        base, suffix = sheet_name, 1
        while sheet_name in existing:
            sheet_name = f"{base[:28]}_{suffix}"
            suffix += 1

        ws = wb.create_sheet(sheet_name)
        write_per_pdf_sheet(ws, rows_verified, pdf_name)

        # Akumulasi untuk sheet GABUNGAN
        for id_val, nilai in rows_verified:
            combined_rows.append((pdf_name, id_val, nilai))

        total_rows_all += len(rows_verified)
        print(f"  ✓ {len(rows_verified)} baris → sheet '{sheet_name}'"
              + (f"  [{n_dibuang} PLU tidak ditemukan, dibuang]" if n_dibuang else ""))

    # Isi sheet GABUNGAN
    write_combined_sheet(ws_combined, combined_rows)

    wb.save(output_path)

    print(f"\n{'='*60}")
    print(f"  Selesai! File Excel tersimpan: {output_path}")
    print(f"  Total baris   : {total_rows_all}")
    print(f"  Sheet         : GABUNGAN + {len(pdf_files)} sheet per-PDF")
    if plu_set is not None:
        print(f"  [PLU DB]      : hanya PLU yang ada di tve8.produk yang masuk Excel")
    print(f"{'='*60}")
    print()
    print("  Tips akurasi OCR:")
    print("  • Gunakan --dpi 400 untuk scan berkualitas rendah.")
    print("  • Pastikan PDF tidak miring / blur.")
    print("  • Gunakan --no-db untuk menonaktifkan filter PLU.")


if __name__ == "__main__":
    main()
